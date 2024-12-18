import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

comodities = "cryptos"

wb = load_workbook(f"final_results_{comodities}.xlsx")
ws = wb.active

width = 26
height = 20
bold_font = Font(bold=True)
title_cells_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
center_alignment = Alignment(horizontal="center", vertical="center")

white = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
green = PatternFill(start_color="51e31b", end_color="51e31b", fill_type="solid")
red = PatternFill(start_color="e31b1b", end_color="e31b1b", fill_type="solid")

# Definovanie ohraničenia
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col in range(1, ws.max_column + 1):
    char = get_column_letter(col)
    ws.column_dimensions[char].width = width

    for row in range(1, ws.max_row + 1):
        cell = ws[f"{char}{row}"]
        ws.row_dimensions[row].height = height
        cell.alignment = center_alignment
        cell.border = thin_border

# Zafarbenie buniek s komoditami
for row in range(1, ws.max_row + 1):
    cell = ws[f"A{row}"]
    cell.font = bold_font
    cell.fill = title_cells_color

# Zafarbenie buniek so stratégiami
for col in range(1, ws.max_column + 1):
    char = get_column_letter(col)
    cell = ws[f"{char}1"]
    cell.font = bold_font
    cell.fill = title_cells_color

# Bunka "comodities" bude biela
first_cell = ws[f"A1"]
first_cell.fill = white

# Zafarbenie buniek s hodnotami podľa úspešnosti
buy_hold_col = ws.max_column - 1

for col in range(2, ws.max_column - 1):
    char = get_column_letter(col)
    for row in range(2, ws.max_row):
        cell = ws[f"{char}{row}"]
        current_value = cell.value
        current_value_formatted = current_value.replace(" ", "").replace(",", ".")
        current_buy_hold_value = ws[f"{get_column_letter(buy_hold_col)}{row}"].value
        current_buy_hold_value_formatted = current_buy_hold_value.replace(" ", "").replace(",", ".")

        if float(current_value_formatted) > float(current_buy_hold_value_formatted):
            cell.fill = green
        if float(current_value_formatted) < 10000:
            cell.fill = red

wb.save(f"final_results_{comodities}.xlsx")

